from calendar import monthcalendar
import datetime
from PIL import Image
from django.conf import settings
from django.core.files.storage import FileSystemStorage

from django.shortcuts import render, redirect, get_object_or_404
from .models import Match, MemberCats, MatchHistory, Members
from django.http import HttpResponse, HttpResponseRedirect, HttpResponseNotFound
from .forms import LoginForm, UserRegistrationForm, ScoreForm, CategoryForm, MembersForm, StartLists
from django.contrib.auth import authenticate, login
from django.contrib.auth.models import User
from django.core.exceptions import ObjectDoesNotExist, MultipleObjectsReturned
from django.db.models import Min, Max
from django.http import JsonResponse
from django.middleware import csrf
import random
from openpyxl import Workbook
from openpyxl.styles import Color, Alignment, Font, PatternFill, Border, Side


points = (
25, 17, 9, 5, 3, 2, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)

cols = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','X','Y','Z']


def set_table(request):
    category_id = request.POST.get('category_id', None)
    match_id = request.POST.get('match_id', None)
    table_number = request.POST.get('table_number', 1)
    user_id = request.user.id
    owner_id = Match.objects.get(pk=match_id).owner_id
    result = None
    if table_number:
        if user_id == owner_id:
            result = MemberCats.objects.filter(pk=category_id).update(table=table_number)
    response = {
        'is_taken': result}
    return JsonResponse(response)
def rollback(request):
    match_id = request.POST.get('match_id', None)
    category_id = request.POST.get('category_id', None)
    user_id = request.user.id
    owner_id = Match.objects.get(pk=match_id).owner_id
    result = None
    if user_id == owner_id:
        step = StartLists.objects.filter(category_id=category_id).aggregate(Max("step"))['step__max']
        members = StartLists.objects.filter(category_id=category_id, step=step)
        for member in members:
            pre_step = StartLists.objects.filter(category_id=category_id, step__lt=step,
                                                 member_id=member.member_id).aggregate(Max("step"))['step__max']
            StartLists.objects.filter(category_id=category_id, member_id=member.member_id,
                                      step=pre_step).update(act=True)
        StartLists.objects.filter(category_id=category_id, step=step).delete()
    response = {
        'is_taken': result}
    return JsonResponse(response)


def set_weight(request):
    match_id = request.POST.get('match_id', None)
    # user_id = int(request.POST.get('user_id', None))
    member_id = request.POST.get('member_id', None)
    user_id = request.user.id
    weight = float(request.POST.get('weight', None))
    owner_id = Match.objects.get(pk=match_id).owner_id
    result = None
    if user_id == owner_id:
        result = Members.objects.filter(pk=member_id).update(weight=weight)
    response = {
        'is_taken': result}
    return JsonResponse(response)


def set_winner(request):
    winner_id = request.POST.get('winner_id', None)
    category_id = request.POST.get('category_id', None)
    match_id = request.POST.get('match_id', None)
    user_id = request.user.id
    owner_id = Match.objects.get(pk=match_id).owner_id
    response = {
        'is_taken': 0}
    if user_id == owner_id:  # Если пользователь владелец матча и может производить эти действия
        winner = StartLists.objects.get(category_id=category_id, member_id=winner_id,
                                        act=True)  # строка спортсмена в StartLists
        loser_id = winner.pair
        loser = StartLists.objects.get(category_id=category_id, member_id=loser_id, act=True)
        position = StartLists.objects.filter(category_id=category_id).aggregate(Max("position"))['position__max']
        place = StartLists.objects.filter(category_id=category_id, place__gt=0).aggregate(Min("place"))['place__min']
        member_count = MemberCats.objects.get(pk=category_id).member_count
        step = StartLists.objects.filter(category_id=category_id).aggregate(Max("step"))['step__max']
        StartLists.objects.filter(category_id=category_id, member_id=winner_id, act=True).update(act=False)
        StartLists.objects.filter(category_id=category_id, member_id=loser_id, act=True).update(act=False)
        if winner.group == 3:  # состоялся поединок за 5 и 6 место
            query = StartLists(category_id=category_id, member_id=winner_id, act=True, position=position + 2,
                               step=step + 1, group=winner.group + 1, match_id=match_id, place=5, fo=loser.fo)
            query.save()
        else:
            query = StartLists(category_id=category_id, member_id=winner_id, act=True, position=position + 1,
                               step=step + 1,
                               group=winner.group, match_id=match_id, fo=winner.fo)  # Победитель на следующую ступень
            query.save()
        if place:
            p = place - 1
        else:
            p = member_count  # определяем текущее место в категории
        if loser.group == 3:  # состоялся поединок за 5 и 6 место
            query = StartLists(category_id=category_id, member_id=loser_id, act=True, position=position + 2,
                               step=step + 1, group=loser.group + 1, match_id=match_id, place=6, fo=loser.fo)
            query.save()
        elif loser.group == 2:  # Если участник выбывает, надо определить его место.
            query = StartLists(category_id=category_id, member_id=loser_id, act=True, position=position + 2,
                               step=step + 1, group=loser.group + 1, match_id=match_id, place=p, fo=loser.fo)
            query.save()
            if p == 2:  # Если проигравший занял второе место то выигравший занял первое
                StartLists.objects.filter(category_id=category_id, member_id=winner_id, act=True).update(place=1)
                MemberCats.objects.filter(pk=category_id).update(final=3)  # Признак завершившейся категории
            elif p == 3:  # значит остался финал
                f = MemberCats.objects.get(pk=category_id).final
                MemberCats.objects.filter(pk=category_id).update(final=f + 1)
        else:
            if p == 2:  # значит остался суперфинал
                MemberCats.objects.filter(pk=category_id).update(final=2)
            query = StartLists(category_id=category_id, member_id=loser_id, act=True, position=position + 2,
                               step=step + 1, group=loser.group + 1, match_id=match_id,
                               fo=loser.fo)  # добавляем без места
            query.save()
    return JsonResponse(response)


def validate_username(request):
    """Проверка доступности логина"""
    username = request.GET.get('username', None)
    response = {
        'is_taken': User.objects.filter(username__iexact=username).exists()
    }

    return JsonResponse(response)


def check_username(request):
    """Проверка наличия спортсмена"""
    s = []
    username = request.POST.get('surname', None)
    category_id = request.POST.get('category_id', None)
    match_id = request.POST.get('match_id', None)
    if len(username) > 2:
        members = Members.objects.filter(surname__icontains=username)
        if members:
            s.append('<div class="popover" id="popover"><ul>')
            for member in members:
                s.append('<li><form action="/dashboard/select_user/" method="POST">')
                s.append('<input type="hidden" name="csrfmiddlewaretoken" value="{}">'.format(csrf.get_token(request)))
                s.append('<input type="hidden" name="match_id" value="{}">'.format(match_id))
                s.append('<input type="hidden" name="category_id" value="{}">'.format(category_id))
                s.append('<input type="hidden" name="member_id" value="{}">'.format(member.pk))
                s.append('<input type="submit" id="go_to" value="{} {} {} {}">'
                         .format(member.surname, member.name, member.second_name, member.birthday))
                s.append('</li>')
            s.append('</ul></div>')
            cal = "\n".join(s)
            response = {
                'is_taken': True,
                'members': cal
            }
        else:
            response = {
                'is_taken': False,
            }
    else:
        response = {
            'is_taken': False,
        }
    return JsonResponse(response)


def month_name(num, lang):
    en = ['january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september',
          'october', 'november', 'december']
    ru = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь', 'июль', 'август', 'сентябрь',
          'октябрь', 'ноябрь', 'декабрь']
    if lang == 'en':
        return en[num - 1]
    else:
        return ru[num - 1]


def get_calendar(date, d):
    s = []
    sm = []
    if date.month == 1:
        pm = 12
        py = date.year - 1
    else:
        pm = date.month - 1
        py = date.year

    if date.month == 12:
        nm = 1
        ny = date.year + 1
    else:
        nm = date.month + 1
        ny = date.year

    pmatrix = monthcalendar(py, pm)
    matrix = monthcalendar(date.year, date.month)
    matrix_c = monthcalendar(date.year, date.month)
    nmatrix = monthcalendar(ny, nm)
    for item in enumerate(pmatrix[-1]):
        if matrix[0][item[0]] == 0:
            matrix[0][item[0]] = item[1]
    for item in enumerate(nmatrix[0]):
        if matrix[-1][item[0]] == 0:
            matrix[-1][item[0]] = item[1]
    m_name = month_name(date.month, 'ru')
    s.append('<table class ="datepicker">\n\t<caption class="datepicker-caption">'
             '\n\t\t<a href="?year={}&month={}" class="datepicker-prev">Previous</a>'.format(py, pm))
    s.append('\t\t<span class="datepicker-title">{} {}</span>'.format(m_name.capitalize(), date.year))
    s.append('\t\t<a href="?year={}&month={}" class="datepicker-next">Next</a>\n\t</caption>\n\t'
             '<thead class="datepicker-head">\n\t\t<tr>\n\t\t\t<th class="datepicker-th">Mo</th>'
             '\n\t\t\t<th class="datepicker-th">Tu</th>\n\t\t\t<th class="datepicker-th">We</th>'
             '\n\t\t\t<th class="datepicker-th">Th</th>\n\t\t\t<th class="datepicker-th">Fr</th>'
             '\n\t\t\t<th class="datepicker-th">Sa</th>\n\t\t\t<th class="datepicker-th">Su</th>'
             '\n\t\t</tr>\n\t</thead>\n\t<tbody class="datepicker-body">'.format(ny, nm))
    for weak in range(len(matrix)):
        s.append('\t\t<tr>')
        for day in range(7):
            if matrix_c[weak][day] == 0:
                tag = 'datepicker-td off'
            else:
                tag = 'datepicker-td'
            if (matrix_c[weak][day] == datetime.datetime.now().day) \
                    and (date.month == datetime.datetime.now().month) and (date.year == datetime.datetime.now().year):
                tag = 'datepicker-td today'
            query = Match.objects.filter(date__year=date.year, date__month=date.month, date__day=matrix_c[weak][day])
            if query:
                tag = 'datepicker-td act'
                for item in query:
                    if d:
                        if int(d) == matrix_c[weak][day]:
                            sm.append('<li class="cart-item">\n\t<span class="cart-item-pic">'
                                      '\t\t<img src="/static/media/{}">'.format(item.poster))
                            sm.append('\t</span><a href="/info/?id={}">{}</a>'.format(item.pk, item.title))
                            sm.append('<span class="cart-item-desc">{}</span>'.format(item.date))
                            sm.append('<span class="cart-item-price">{}</span>'.format(item.location))
                            sm.append('</li>')
                            break
                    else:
                        sm.append('<li class="cart-item">\n\t<span class="cart-item-pic">'
                                  '\t\t<img src="/static/media/{}">'.format(item.poster))
                        sm.append('\t</span><a href="/info/?id={}">{}</a>'.format(item.pk, item.title))
                        sm.append('<span class="cart-item-desc">{}</span>'.format(item.date))
                        sm.append('<span class="cart-item-price">{}</span>'.format(item.location))
                        sm.append('</li>')
            s.append('\t\t\t<td class="{0}"><a href="?year={1}&month={2}&day={3}">{3}</a></td>'.format(tag, date.year,
                                                                                                       date.month,
                                                                                                       matrix[weak][
                                                                                                           day]))
        s.append('\t\t</tr>')
    s.append('\t</tbody>\n</table>')
    cal = "\n".join(s)
    lst = "\n".join(sm)
    return cal, lst


def get_matches(user_id):
    matches = Match.objects.filter(owner_id=user_id)
    sm = []
    for item in matches:
        sm.append('<li class="cart-item">\n\t<span class="cart-item-pic">'
                  '\t\t<img src="/static/media/{}">'.format(item.poster))
        sm.append('\t</span><a href="/info/?id={}">{}</a>'.format(item.pk, item.title))
        sm.append('<span class="cart-item-desc">{}</span>'.format(item.date))
        sm.append('<span class="cart-item-price">{}</span>'.format(item.location))
        sm.append('</li>')
    lst = "\n".join(sm)
    return lst


def get_menu(request):
    auth = request.user.is_authenticated
    s = []
    if auth:
        s.append('<a href="/accounts/logout/" class="menu_button">Выйти</a>')
        s.append('<a href="/dashboard/matches/" class="menu_button">Мои турниры</a>')
        s.append('<form action="/dashboard/match/" method="post">')
        s.append('<input type="hidden" name="csrfmiddlewaretoken" value="{}">'.format(csrf.get_token(request)))
        s.append('<input type="hidden" name="selector" value="add_match">')
        s.append('<input type="submit" class="btn" id="go_to_match" value="Создать турнир">')
        s.append('</form>')
    else:
        s.append('<a href="/accounts/login/" class="menu_button">Войти</a>')
    # s.append('<a href="/accounts/login/" class="menu_button">Рейтинги</a>')
    s.append('<a href="/members/" class="menu_button">Спортсмены</a>')
    # s.append('<a href="/accounts/login/" class="menu_button">Организаторы</a>')
    # s.append('<a href="/accounts/login/" class="menu_button">Клубы</a>')
    s.append('<a href="/" class="menu_button">Главная</a>')
    # s.append('')
    result = "\n".join(s)
    return result


def get_menu_cat(match_id, request):
    user_id = request.user.id
    owner_id = Match.objects.get(pk=match_id).owner_id
    s = []
    if user_id == owner_id:
        s.append('<form method="post" action="/dashboard/match/">')
        s.append('<input type="hidden" name="csrfmiddlewaretoken" value="{}">'.format(csrf.get_token(request)))
        s.append('<input type="hidden" name="selector" value="add_category" />')
        s.append('<input type="hidden" name="match_id" value={} />'.format(match_id))
        s.append('<input type="submit" id="go_to_match"  value="Добавить категорию" />')
        s.append('</form>')
        s.append('<form method="post" action="/dashboard/match/">')
        s.append('<input type="hidden" name="csrfmiddlewaretoken" value="{}">'.format(csrf.get_token(request)))
        s.append('<input type="hidden" name="selector" value="start_match" />')
        s.append('<input type="hidden" name="match_id" value={} />'.format(match_id))
        s.append('<input type="submit" id="go_to_match"  value="Турнир" />')
        s.append('</form>')
        s.append('</form>')
    s.append('<form method="post" action="/dashboard/match/">')
    s.append('<input type="hidden" name="csrfmiddlewaretoken" value="{}">'.format(csrf.get_token(request)))
    s.append('<input type="hidden" name="selector" value="start_lists" />')
    s.append('<input type="hidden" name="match_id" value={} />'.format(match_id))
    s.append('<input type="submit" id="go_to_match"  value="Стартовые списки" />')
    s.append('</form>')
    s.append('<form method="post" action="/dashboard/match/">')
    s.append('<input type="hidden" name="csrfmiddlewaretoken" value="{}">'.format(csrf.get_token(request)))
    s.append('<input type="hidden" name="selector" value="table_list" />')
    s.append('<input type="hidden" name="match_id" value={} />'.format(match_id))
    s.append('<input type="submit" id="go_to_match"  value="Табло онлайн" />')
    s.append('</form>')
    result = "\n".join(s)
    return result


def get_tables(user_id, match_id, token):  # Список проведения турнира (Турнир)
    owner_id = Match.objects.get(pk=match_id).owner_id
    cats = MemberCats.objects.filter(match_id=match_id).order_by("hand", "table")
    s = []
    counter_name = 0  # Идентификатор для определения значения
    for category in cats:  # Для каждой категории
        if category.started:
            counter_name = counter_name + 1
            # Определяем пары
            members = StartLists.objects.filter(category_id=category.pk, act=True).order_by("place", "position")
            count = StartLists.objects.filter(category_id=category.pk, act=True,
                                              group=1).count()  # определяем финалиста 1
            count2 = StartLists.objects.filter(category_id=category.pk, act=True, group=2).count()  # определяем финал
            c56 = Match.objects.get(pk=match_id).c56
            member_list = []
            result_list = []
            last_members = []
            state = 0
            i = 0  #
            for member in members:  # Список участников (чтобы можно было применить метод pop для удаления участника)
                member_list.append(member)
            while member_list:
                if member_list[i].place > 0:  # Список выбывших из соревнования
                    # Если в списке встречаем 5 а следующее 6 место и надо провести дополнительный поединок
                    if member_list[i].place == 5 and c56 and member_list[i].group == 3:
                        result_list.append(member_list.pop(i))  # добавляем 5 место
                        result_list.append(member_list.pop(i))  # добавляем 6 место
                        StartLists.objects.filter(category_id=category.pk, act=True,  # Пары в базе
                                                  pk=result_list[-1].pk).update(pair=result_list[-2].member_id)
                        StartLists.objects.filter(category_id=category.pk, act=True,
                                                  pk=result_list[-2].pk).update(pair=result_list[-1].member_id)
                        result_list[-1].pair = result_list[-2].member_id
                        result_list[-1].place = 0
                        result_list[-2].pair = result_list[-1].member_id
                        result_list[-2].place = 0
                        continue
                    else:
                        last_members.insert(0, member_list.pop(i))
                    i = 0
                    state = 0  # если дошли до конца списка, значит пары нет и можно искать следующую пару
                    continue

                if state == 0:  # если первый участник пары
                    if count == 1 and member_list[i].group == 1:  # Если остался один участник в А группе
                        last_members.append(member_list.pop(i))  # Первый финалист
                        if count2 == 1:
                            state = 1  # если финал, то ищем пару финалисту.
                    else:  # иначе просто добавляем участника в новую пару
                        result_list.append(member_list.pop(i))
                        state = 1  # Флаг поиска пары
                    i = 0
                    continue
                elif state == 1:
                    if count == 1 and count2 == 1:  # Если финал
                        last_members.append(member_list.pop(i))
                        StartLists.objects.filter(category_id=category.pk, act=True,  # Пары в базе
                                                  pk=last_members[-1].pk).update(pair=last_members[-2].member_id)
                        StartLists.objects.filter(category_id=category.pk, act=True,
                                                  pk=last_members[-2].pk).update(pair=last_members[-1].member_id)
                        last_members[-2].pair = last_members[-1].member_id
                        last_members[-1].pair = last_members[-2].member_id
                        continue
                    if result_list[-1].group == member_list[i].group:  # является парой?
                        result_list.append(member_list.pop(i))  # Если нашли пару в группе, добавляем следующим
                        StartLists.objects.filter(category_id=category.pk, act=True,  # Пары в базе
                                                  pk=result_list[-1].pk).update(pair=result_list[-2].member_id)
                        StartLists.objects.filter(category_id=category.pk, act=True,
                                                  pk=result_list[-2].pk).update(pair=result_list[-1].member_id)
                        result_list[-1].pair = result_list[-2].member_id
                        result_list[-2].pair = result_list[-1].member_id
                        state = 0
                        i = 0
                    else:  # если не пара то
                        if len(member_list) - 1 == i:  # Последнего участника просто добавляем
                            result_list.append(member_list.pop(i))
                        else:
                            i = i + 1

            if last_members:
                last_members = sorted(last_members, key=lambda lists: lists.place)
                result_list.extend(last_members)
            if category.hand == '1':
                h = 'правая рука'
            else:
                h = 'левая рука'
            s.append('<div class="add_category">')
            s.append('<table class="tab_group">\n\t<tbody>\n\t\t<tr class="row_cat">\n\t\t\t<th colspan="3">'
                     'Стол: {} Категория: {} {} {} ({})</th>'.format(category.table, category.age_category,
                                                                     category.weight_category,
                                                                     category.group_category, h))
            if user_id == owner_id and category.final < 3:
                s.append('<th class="coll_field weight-input"><form>'.format(counter_name))
                s.append('<input type="hidden" id="csrfmiddlewaretoken" name="csrfmiddlewaretoken" value="{}">'.format(token))
                s.append('<input type="hidden" id="category_id_{0}" name="category_id_{0}" value="{0}" />'.format(category.pk))
                s.append('<input type="hidden" id="match_id_{0}" name="match_id_{0}" value="{1}" />'.format(category.pk, match_id))
                s.append('<input type="number" id="table_number_{0}" class="small_number" name="table_number_{0}">'.format(category.pk))
                s.append('<button id="{}" name="set_table" type="submit" class="small_button">Изм. № стола</button></form></th>'.format(category.pk))
            else:
                s.append('<th></th>')
            s.append('</tr>')
            if category.final == 0:
                s.append('<tr class="row_cat"><th colspan="4" class="coll_field">Следующий поединок</th></tr>')
            elif category.final == 1:
                s.append('<tr class="row_cat"><th colspan="4" class="coll_field">Финал</th></tr>')
            elif category.final == 2:
                s.append('<tr class="row_cat"><th colspan="4" class="coll_field">Суперфинал</th></tr>')
            elif category.final == 3:
                s.append('<tr class="row_cat"><th colspan="4" class="coll_field">Результаты</th></tr>')
            if user_id == owner_id:
                if members.count() > 0:  # Если в категории есть участники
                    if result_list[0].pair > 0:  # Если в категории есть пары участников
                        s.append('<tr class="row_cat">')
                        s.append('\t\t\t<td class="coll_min">')
                        s.append('<form method="post">')
                        s.append(
                            '<input type="hidden" id="csrfmiddlewaretoken" name="csrfmiddlewaretoken" value="{}">'.format(
                                token))
                        s.append('<input type="hidden" id="winner_id_{0}_{1}" name="winner_id_{0}_{1}" value={0} />'.format(
                            result_list[0].member_id, category.pk))
                        s.append(
                            '<input type="hidden" id="category_id_{0}_{1}" name="category_id_{0}_{1}" value={1} />'.format(result_list[0].member_id, category.pk))
                        s.append('<input type="hidden" id="match_id_{0}_{1}" name="match_id_{0}_{1}" value={2} />'.format(result_list[0].member_id, category.pk, match_id))
                        s.append('<button id="{}_{}" name="set_winner" type="submit" class="small_button">Победил</button>'.format(result_list[0].member_id, category.pk))
                        s.append('</form>')
                        s.append('\t\t\t</td>')
                        if category.started:
                            s.append('\t\t\t<td class="coll_button">')
                            s.append(
                                '<form method="post">')
                            s.append(
                                '<input type="hidden" id="csrfmiddlewaretoken" name="csrfmiddlewaretoken" value="{}">'.format(
                                    token))
                            s.append(
                                '<input type="hidden" id="category_id_{0}_{1}" name="category_id_{0}_{1}" value={1} />'.format(
                                    match_id, category.pk))
                            s.append(
                                '<input type="hidden" id="match_id_{0}_{1}" name="match_id_{0}_{1}" value={0} />'.format(
                                    match_id, category.pk))
                            s.append(
                                '<button id="{0}_{1}" name="rollback" type="submit" class="small_button">Отменить</button>'.format(
                                    match_id, category.pk))
                            s.append('</form>')
                            s.append('\t\t\t</td>')
                            s.append('<td></td>')
                        else:
                            s.append('<td colspan="2"></td>')
                        s.append('\t\t\t<td class="coll_min">')
                        s.append('<form method="post">')
                        s.append(
                            '<input type="hidden" id="csrfmiddlewaretoken" name="csrfmiddlewaretoken" value="{}">'.format(
                                token))
                        s.append(
                            '<input type="hidden" id="winner_id_{0}_{1}" name="winner_id_{0}_{1}" value={0} />'.format(
                                result_list[1].member_id, category.pk))
                        s.append(
                            '<input type="hidden" id="category_id_{0}_{1}" name="category_id_{0}_{1}" value={1} />'.format(
                                result_list[1].member_id, category.pk))
                        s.append(
                            '<input type="hidden" id="match_id_{0}_{1}" name="match_id_{0}_{1}" value={2} />'.format(
                                result_list[1].member_id, category.pk, match_id))
                        s.append(
                            '<button id="{}_{}" name="set_winner" type="submit" class="small_button">Победил</button>'.format(
                                result_list[1].member_id, category.pk))
                        s.append('</form>')
            s.append('\t\t\t</td></tr>')
            s.append('</tbody></table></div>')
            s.append('<div class="add_category">')
            s.append('<table class="tab_group">\n\t<tbody>')
            k = 0  # какой в паре
            for i in range(len(result_list)):
                if result_list[i]:
                    m = Members.objects.get(pk=result_list[i].member_id)
                    if result_list[i].place == 0:  # Если еще не вылетел
                        if result_list[i].pair > 0:  # Если у участника есть пара
                            if k == 0:  # Если первый в паре
                                s.append('<tr class="row_member">')
                                if result_list[i].group == 1:
                                    s.append('<td class="coll_min">А</td>')
                                if result_list[i].group == 2:
                                    s.append('<td class="coll_min">B</td>')
                                if result_list[i].group == 3:
                                    if result_list[i].place == 0:
                                        s.append('<td class="coll_min">5 или 6</td>')
                                if result_list[i].pair > 0:
                                    s.append(
                                        '<td class="coll_name">{} {} {}</td>'.format(m.surname, m.name, m.second_name))
                                k = 1
                            else:  # Если второй в паре
                                if result_list[i].pair > 0:
                                    s.append('<td class="coll_name">{} {} {}</td></tr>'.format(m.surname, m.name,
                                                                                               m.second_name))
                                k = 0
                        else:  # Если у участника нет пары
                            s.append('<tr class="row_member">')
                            if result_list[i].group == 1:
                                s.append('<td class="coll_min">А</td>')
                            if result_list[i].group == 2:
                                s.append('<td class="coll_min">B</td>')
                            s.append('<td class="coll_name" colspan="3">{} {} {}</td>'.format(m.surname, m.name,
                                                                                              m.second_name))
                            s.append('</tr>')
                    else:  # Если участник уже занял какое-то место
                        s.append('<tr class="row_member">')
                        s.append('<td class="coll_min">{} местo</td>'.format(result_list[i].place))
                        s.append(
                            '<td class="coll_name" colspan="2">{} {} {}</td>'.format(m.surname, m.name, m.second_name))
                        s.append('</tr>')
                else:
                    s.append('<tr class="row_member"><td class="coll_min"></td><td class="coll_name"></td>')

            s.append('</tbody></table></div>')
    lst = "\n".join(s)
    return lst


def get_list(match_id, request, token):  # Стартовые списки
    user_id = request.user.id
    owner_id = Match.objects.get(pk=match_id).owner_id
    s = []
    categories = MemberCats.objects.filter(match_id=match_id)
    for category in categories:
        members = StartLists.objects.filter(category_id=category.pk, act=True).order_by('place')
        if category.hand == '1':
            hand = 'правая рука'
        else:
            hand = 'левая рука'
        s.append('<div class="add_category">')
        s.append('<table class="tab_group">\n\t<tbody>\n\t\t<tr class="row_cat">\n\t\t\t<th colspan="3">'
                 'Категория: {} {} {} ({})</th>'.format(category.age_category, category.weight_category,
                                                        category.group_category, hand))
        if user_id == owner_id:
            if category.final < 3:  # Если категория еще не завершена ее можно остановить или начать
                if not category.started:
                    if category.member_count > 1:
                        s.append('\t\t\t<td class="coll_button">')
                        s.append('<form method="post" action="/dashboard/match/">')
                        s.append('<input type="hidden" name="csrfmiddlewaretoken" value="{}">'.format(token))
                        s.append('<input type="hidden" name="selector" value="start_category" />')
                        s.append('<input type="hidden" name="category_id" value={} />'.format(category.pk))
                        s.append('<input type="hidden" name="match_id" value={} />'.format(match_id))
                        s.append('<input type="submit" id="row_cat"  value="Начать категорию" />')
                        s.append('</form>\n\t\t\t</td>')
                else:
                    s.append('\t\t\t<td class="coll_button">')
                    s.append('<form method="post" action="/dashboard/match/">')
                    s.append('<input type="hidden" name="csrfmiddlewaretoken" value="{}">'.format(token))
                    s.append('<input type="hidden" name="selector" value="stop_category" />')
                    s.append('<input type="hidden" name="category_id" value={} />'.format(category.pk))
                    s.append('<input type="hidden" name="match_id" value={} />'.format(match_id))
                    s.append('<input type="submit" id="row_cat"  value="Остановить категорию" />')
                    s.append('</form>\n\t\t\t</td>')
            else:  # добавим кнопку вывода результатов в PDF
                s.append('\t\t\t<td class="coll_button">')
                s.append('<form method="post" action="/dashboard/match/">')
                s.append('<input type="hidden" name="csrfmiddlewaretoken" value="{}">'.format(token))
                s.append('<input type="hidden" name="selector" value="print_result" />')
                s.append('<input type="hidden" name="category_id" value={} />'.format(category.pk))
                s.append('<input type="hidden" name="match_id" value={} />'.format(match_id))
                s.append('<input type="submit" id="row_cat"  value="Отчет XLSX" />')
                s.append('</form>\n\t\t\t</td>')
            if not category.started:
                s.append('\t\t\t<td class="coll_button">')
                s.append('<form method="post" action="/dashboard/match/">')
                s.append('<input type="hidden" name="csrfmiddlewaretoken" value="{}">'.format(token))
                s.append('<input type="hidden" name="selector" value="delete_category" />')
                s.append('<input type="hidden" name="category_id" value={} />'.format(category.pk))
                s.append('<input type="hidden" name="match_id" value={} />'.format(match_id))
                s.append('<input type="submit" id="row_cat"  value="Удалить категорию" />')
                s.append('</form>\n\t\t\t</td>')

                s.append('\t\t\t<td class="coll_button">')
                s.append('<form method="post" action="/dashboard/match/">')
                s.append('<input type="hidden" name="csrfmiddlewaretoken" value="{}">'.format(token))
                s.append('<input type="hidden" name="selector" value="add_member" />')
                s.append('<input type="hidden" name="category_id" value={} />'.format(category.pk))
                s.append('<input type="hidden" name="match_id" value={} />'.format(match_id))
                s.append('<input type="submit" id="row_cat"  value="Добавить участника" />')
                s.append('</form>\n\t\t\t</td>')
            else:
                s.append('\t\t\t<td class="coll_button"></td><td class="coll_button"></td>')
        s.append('</tr>')
        s.append('</tbody></table></div>')
        s.append('<table class="tab_group">\n\t<tbody>')
        s.append('<tr class="row_header">')
        s.append('\t<th class="coll_field">Фамилия, Имя, Отчество</th>')
        s.append('\t<th class="coll_field">Спортивное звание</th>')
        s.append('\t<th class="coll_field">Дата рождения</th>')
        s.append('\t<th class="coll_field">Федеральный округ</th>')
        s.append('\t<th class="coll_field">Команда</th>')
        if category.final == 3:
            s.append('\t<th class="coll_field">Вес</th>')
            s.append('\t<th class="coll_field">Место</th>')
            s.append('\t<th class="coll_field">Очки</th>')
        else:
            s.append('\t<th class="coll_field" colspan=3>Вес</th>')
        s.append('</tr>')
        for member in members:
            try:  # есть ли спортсмен
                m = Members.objects.get(pk=member.member_id)
            except ObjectDoesNotExist:  # если спортсмен был удален (но остался в списках, то удаляем его из списков)
                StartLists.objects.filter(member_id=member.member_id).delete()
                continue
            s.append('<tr class="row_member">')
            s.append('\t\t\t<td class="coll_name">{} {} {}</td>'.format(m.surname, m.name, m.second_name))
            s.append('\t\t\t<th class="coll_field">{}</th>'.format(m.rank))
            if m.birthday:
                s.append('\t\t\t<th class="coll_field">{}</th>'.format(m.birthday))
            else:
                s.append('\t\t\t<td class="coll_field"></td>')
            s.append('\t\t\t<th class="coll_name">{}</th>'.format(m.fo))
            s.append('\t\t\t<th class="coll_field">{}</th>'.format(m.team))
            # Измерение веса участника
            if user_id == owner_id and not category.started:
                s.append('<td class="coll_field weight-input">'
                         '<form method="post" action="/dashboard/match/" id="set_weight">'
                         '<input type="number" id="weight_{1}_{3}" name="weight_{1}_{3}" placeholder="{2}" value="{2}" step="0.01" min="0">'
                         '<input type="hidden" id="match_id_{1}_{3}" name="match_id_{1}_{3}" value="{0}">'
                         '<input type="hidden" id="category_id_{1}" name="category_id_{1}" value="{3}">'
                         '<input type="hidden" id="member_id_{1}_{3}" name="member_id_{1}_{3}" value="{1}">'.format(match_id, m.pk, m.weight, category.pk))
                s.append('<input type="hidden" id="csrfmiddlewaretoken" name="csrfmiddlewaretoken" value="{}">'.format(
                    token))
                # s.append('<button id="{}" type="submit" class="button">Сохранить</button>'.format(m.pk))
                s.append('</form></td>')
                s.append('<td><button id="{}_{}" name="set_weight" type="submit" class="button">Сохранить</button></td>'.format(m.pk, category.pk))
                s.append('\t\t\t<td class="coll_button">')
                s.append('<form method="post" action="/dashboard/match/">')
                s.append('<input type="hidden" name="csrfmiddlewaretoken" value="{}">'.format(token))
                s.append('<input type="hidden" name="selector" value="delete_member" />')
                s.append('<input type="hidden" name="member_id" value={} />'.format(member.member_id))
                s.append('<input type="hidden" name="category_id" value={} />'.format(category.pk))
                s.append('<input type="hidden" name="match_id" value={} />'.format(match_id))
                s.append('<input type="submit" id="row_cat"  value="Удалить участника" />')
                s.append('</form>\n\t\t\t</td>')
            else:
                if category.final == 3:
                    s.append('\t\t\t<th class="coll_field">{}</th>'.format(m.weight))
                    s.append('\t\t\t<th class="coll_field" >{}</th>'.format(member.place))
                    s.append('\t\t\t<th class="coll_field" >{}</th>'.format(points[member.place - 1]))
                else:
                    s.append('\t\t\t<th class="coll_field colspan="3">{}</th>'.format(m.weight))
            s.append('</tr>')
        s.append('</tbody></table>')
    result = "\n".join(s)
    return result


def get_header(user):
    s = []
    if user:
        s.append('<div class="header-login">Вы вошли как {0}</div>'.format(user))
    s.append('<div class="header-text">Платформа проведения турниров по армрестлингу</div>')
    result = "\n".join(s)
    return result


def get_match_info(match_id, user_id, request):
    s = []
    item = Match.objects.get(pk=match_id)
    s.append('<div class=cart>')
    s.append('\t\t<div class="cart-top"><h2 class="cart-top-title">{}</h2></div>'.format(item.title))
    s.append('<ul>\n<li class="cart-item"><span class="poster-item-pic">'
             '<img src="/static/media/{}">\n</span>'.format(item.poster))
    s.append('<span class="cart-item-desc">Дата проведения: {}</span>'.format(item.date))
    s.append('<span class="cart-item-desc">Регион: {}</span>'.format(item.fo))
    s.append('<span class="cart-item-desc">Место проведения: {}</span>'.format(item.location))
    s.append('<span class="cart-item-desc">Главный судья: {}</span>'.format(item.gj))
    s.append('<span class="cart-item-desc">Главный секретарь: {}</span>'.format(item.gs))
    if item.hands == '0':
        hands = 'Борьба на обоих руках'
    elif item.hands == '1':
        hands = 'Борьба на правой руке'
    else:
        hands = 'Борьба на левой руке'
    s.append('<span class="cart-item-desc">{}</span>'.format(hands))
    s.append('<span class="cart-item-desc">Количество столов: {}</span>'.format(item.table_count))
    s.append('<span class="cart-item-desc">Контакты: {}</span>'.format(item.contacts))
    s.append('<span class="cart-item-desc">Положение турнира: {}</span>'.format(item.gs))
    s.append('</li>\n</ul>')
    s.append('<form method="post" action="/dashboard/match/">')
    s.append('<input type="hidden" name="csrfmiddlewaretoken" value="{}">'.format(csrf.get_token(request)))
    s.append('<input type="hidden" name="match_id" value="{}">'.format(match_id))
    s.append('<input type="hidden" name="selector" value="start_lists">')
    s.append('<input type="submit" id="go_to_match" value="Стартовые списки">\n</form>')
    if item.owner_id == user_id:
        s.append('<form method="post" action="/dashboard/match/">')
        s.append('<input type="hidden" name="csrfmiddlewaretoken" value="{}">'.format(csrf.get_token(request)))
        s.append('<input type="hidden" name="selector" value="edit_match">')
        s.append('<input type="hidden" name="match_id" value="{}">'.format(match_id))
        s.append('<input type="submit" id="go_to_match" value="Редактировать">\n</form>')
        s.append('<form method="post" action="/dashboard/match/">')
        s.append('<input type="hidden" name="csrfmiddlewaretoken" value="{}">'.format(csrf.get_token(request)))
        s.append('<input type="hidden" name="match_id" value="{}">'.format(match_id))
        s.append('<input type="hidden" name="selector" value="delete_match">')
        s.append('<input type="submit" id="go_to_match" value="Удалить">\n</form>')
    s.append('</div>')

    match = "\n".join(s)
    return match


def get_annonce(date):  # Импорт из ВК
    s = []
    s.append('<script src="https://vk.com/js/api/openapi.js?168" type="text/javascript"></script>')
    s.append('<table class="annonce"><tr><th><div id="vk_groups"></div></th></tr></table><script type="text/javascript">VK.Widgets.Group("vk_groups",'
             ' {mode: 4, wide: 1, width: 600, height: 800,  color1: "242930", color2: "8a99af", color3: "8a99af"},'
             ' 206407744);</script>')
    result = "\n".join(s)
    return result



def index(request):  # точка входа в приложение, по результатам авторизации
    data = {}
    date = datetime.datetime.now()
    month = date.month
    user = None
    day = None
    if request.method == 'GET':
        month = request.GET.get("month", "")
        day = request.GET.get("day", "")
        year = request.GET.get("year", "")
        if year:
            if day:
                date = datetime.date.fromisoformat('{0}-{1:0>2}-{2:0>2}'.format(year, month, day))
            else:
                date = datetime.date.fromisoformat('{0}-{1:0>2}-{2:0>2}'.format(year, month, 1))
    if request.user.is_authenticated == True:
        user = request.user
        data['username'] = user
    data['header'] = get_header(user)
    cal, lst = get_calendar(date, day)
    data['calendar'] = cal
    data['matches'] = lst
    data['menu'] = get_menu(request)
    data['annonce'] = get_annonce(datetime.datetime.now())
    if month:
        data['month_name'] = 'Турниры за {}'.format(month_name(int(month), 'ru'))
    return render(request, "pages/index.html", context=data)


def info(request):
    data = {}
    user_id = None
    if request.user.is_authenticated == True:
        user = request.user
        user_id = request.user.id
        data['username'] = user
        data['header'] = get_header(user)
    data['menu'] = get_menu(request)
    if request.method == 'GET':
        match_id = request.GET.get("id", "")
        if match_id:
            match = get_match_info(match_id, user_id, request)
            data['match'] = match
            return render(request, "pages/match_info.html", context=data)
        else:
            return redirect("/")


def user_login(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            user = authenticate(username=cd['username'], password=cd['password'])
            if user is not None:
                if user.is_active:
                    login(request, user)
                    user_id = request.user.id
                    # matches = Match.objects.filter(owner_id=user_id, arch=0)
                    # return render(request, 'pages/dashboard.html', {"username": user, 'matches': matches})
                    return redirect("/")
                else:
                    return HttpResponse('Disabled pages')
            else:
                return HttpResponse('Invalid login')
    else:
        form = LoginForm()
    return render(request, 'registration/login.html', {'form': form})


def user_logout(request):
    return render(request, 'registration/logged_out.html')


def user_password_change(request):
    return render(request, 'registration/password_change_form.html')


def user_password_change_done(request):
    return render(request, 'registration/password_change_done.html')


def user_password_reset(request):
    return render(request, 'registration/password_reset_form.html')


def user_password_reset_done(request):
    return render(request, 'registration/password_reset_done.html')


def password_reset_confirm(request):
    return render(request, 'registration/password_reset_confirm.html')


def password_reset_done(request):
    return render(request, 'registration/password_reset_done.html')


def user_registration(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            new_user = form.save(commit=False)
            new_user.set_password(form.cleaned_data['password'])
            new_user.save()
            return render(request, 'registration/login.html', {})
    else:
        form = UserRegistrationForm()
    return render(request, 'registration/register.html', {'form': form})


def show_match_list(request):
    if request.user.is_authenticated:
        user = request.user
        data = {}
        if user:
            if user.is_active:
                login(request, user)
                user_id = request.user.id
                data['header'] = get_header(user)
                data['menu'] = get_menu(request)
                data['page_name'] = 'Мои турниры'
                data['matches'] = get_matches(user_id)
            return render(request, 'pages/dashboard.html', context=data)
    return render(request, 'pages/index.html', {})


def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def match(request):
    data = {}
    user_id = None
    user = None
    match_id = None
    if request.user.is_authenticated:
        user = request.user
        user_id = request.user.id
    data['header'] = get_header(user)
    data['menu'] = get_menu(request)

    if request.method == 'GET':  # редирект
        match_id = request.GET.get("id")
        if match_id:
            data['start_list'] = get_list(match_id, request, csrf.get_token(request))
            data['menu_cat'] = get_menu_cat(match_id, request)
            query = Match.objects.get(pk=match_id)
            data['title'] = query.title
    elif request.method == 'POST':
        # form = ScoreForm(request.POST) # заполнить форму данными
        action = request.POST.get("selector", "Undefined")
        match_id = request.POST.get("match_id")
        if match_id:
            query = Match.objects.get(pk=match_id)
            data['title'] = query.title
            data['date'] = query.date
            data['menu_cat'] = get_menu_cat(match_id, request)
        else:  # если редирект то возвращаем стартовые списки
            pass
        if action == 'edit_match':
            match = Match.objects.get(pk=match_id)
            form = ScoreForm(initial={'title': match.title, 'location': match.location, 'fo': match.fo,
                                      'date': '{0}-{1:0>2}-{2:0>2}'.format(match.date.year, match.date.month,
                                                                           match.date.day),
                                      'gj': match.gj, 'gs': match.gs, 'table_count': match.table_count,
                                      'contacts': match.contacts, 'hands': match.hands, 'poster': match.poster,
                                      'c56': match.c56})
            form.match_id = match.pk
            form.owner_id = match.owner_id
            data['edit_form'] = form
            data['match_id'] = match_id
            return render(request, 'pages/edit_match.html', context=data)

        if action == 'update_match':
            user_id = request.user.id
            instance = get_object_or_404(Match, id=match_id)
            form = ScoreForm(request.POST or None, request.FILES, instance=instance)
            if form.is_valid():
                form.match_id = match_id
                form.owner_id = user_id
                form.save()
            return redirect('/info/?id={}'.format(match_id))

        if action == 'add_match':
            form = ScoreForm(request.POST or None, request.FILES)
            data['score_form'] = form
            return render(request, 'pages/match.html', context=data)

        if action == 'set_match':
            user_id = request.user.id  # id пользователя организатора соревнования
            form = ScoreForm(request.POST or None, request.FILES)
            if form.is_valid():
                form.owner_id = user_id
                last = form.save()
                return redirect('/info/?id={}'.format(last.pk))
            return redirect(request.path)

        if action == 'reset_match':  # сброс турнира
            if match_id:
                MemberCats.objects.filter(match_id=match_id).update(status=0)
                StartLists.objects.filter(match_id=match_id).update(place=0, position=0, step=0, act=False,
                                                                    group=1, pair=None)
                data['start_list'] = get_list(match_id, request, csrf.get_token(request))

        if action == 'delete_match':  # удаление турнира
            if match_id:
                Match.objects.filter(id=match_id).delete()
                MatchHistory.objects.filter(match=match_id).delete()
            return redirect('/')

        if action == 'add_category':  # добавление категории
            form = CategoryForm()
            data['category_form'] = form

        if action == 'set_category':
            form = CategoryForm(request.POST, request.FILES)
            form.match_id = match_id
            if form.is_valid():
                form.save()
            data['start_list'] = get_list(match_id, request, csrf.get_token(request))
            return redirect("/dashboard/match/?id={}".format(match_id))

        if action == 'delete_category':  # удаление категории
            category_id = request.POST.get('category_id')
            MemberCats.objects.filter(pk=category_id).delete()
            data['start_list'] = get_list(match_id, request, csrf.get_token(request))
            MatchHistory.objects.filter(category=category_id).delete()

        if action == 'start_category':  # Начало борьбы в категории
            category_id = request.POST.get('category_id')
            match_id = request.POST.get('match_id')
            list_id = list(
                StartLists.objects.filter(category_id=category_id).values_list("id",
                                                                               flat=True))  # id участников в категории
            random.shuffle(list_id)  # перемешиваем для жеребьевки
            MemberCats.objects.filter(pk=category_id).update(member_count=len(list_id))
            position = 1
            while list_id:  # жеребьевка для категории, создание очереди
                member_id = list_id.pop(0)
                StartLists.objects.filter(pk=member_id).update(position=position, step=0, act=True, group=1, pair=0,
                                                               place=0)
                position = position + 1
            MemberCats.objects.filter(pk=category_id).update(started=True)
            data['start_list'] = get_list(match_id, request, csrf.get_token(request))

        if action == 'stop_category':  # Сброс борьбы в категории
            category_id = request.POST.get('category_id')
            StartLists.objects.filter(category_id=category_id, step__gte=1).delete()
            StartLists.objects.filter(category_id=category_id).update(position=0, place=0, group=1, act=True, pair=0)
            MemberCats.objects.filter(pk=category_id).update(started=False, final=False)
            data['start_list'] = get_list(match_id, request, csrf.get_token(request))

        if action == 'add_member':
            form = MembersForm()
            data['category_id'] = request.POST.get('category_id')
            data['match_id'] = request.POST.get('match_id')
            data['members_form'] = form
            cat = MemberCats.objects.get(pk=data['category_id'])
            if not cat.started:
                return render(request, 'pages/add_member.html', context=data)

        if action == 'set_member':
            category_id = request.POST.get('category_id')
            form = MembersForm(request.POST, request.FILES)
            form.match_id = match_id
            form.category_id = category_id
            if form.is_valid():
                just_created = form.save()
                StartLists.objects.create(member_id=just_created.id, category_id=category_id, match_id=match_id,
                                          fo=just_created.fo)
                count = StartLists.objects.filter(category_id=category_id, step=0).count()
                MemberCats.objects.filter(pk=category_id).update(member_count=count)
            data['start_list'] = get_list(match_id, request, csrf.get_token(request))

        if action == 'start_lists':
            data['start_list'] = get_list(match_id, request, csrf.get_token(request))

        if action == 'print_result':
            wb = Workbook()
            ws = wb.active
            ws.title = "Результаты турнира"
            # Формирование титула
            m1 = Match.objects.get(pk=match_id)
            ws.merge_cells('{}{}:{}{}'.format(cols[1], 1, cols[8], 1))
            cell = '{}{}'.format(cols[1], 1)
            ws[cell].value = 'Турнир: ' + m1.title
            ws[cell].font = Font(bold=True)
            ws[cell].alignment = Alignment(horizontal='center')
            ws.merge_cells('{}{}:{}{}'.format(cols[1], 2, cols[8], 2))
            cell = '{}{}'.format(cols[1], 2)
            ws[cell].value = 'Место проведения: ' + m1.location
            ws[cell].font = Font(bold=True)
            ws[cell].alignment = Alignment(horizontal='center')
            ws.merge_cells('{}{}:{}{}'.format(cols[1], 3, cols[8], 3))
            cell = '{}{}'.format(cols[1], 3)
            ws[cell].value = 'Дата проведения: {}'.format(m1.date)
            ws[cell].font = Font(bold=True)
            ws[cell].alignment = Alignment(horizontal='center')
            query = StartLists.objects.filter(match_id=match_id,  act=True).order_by('category_id', 'place')
            # Формирование заголовка отчета для категории
            row = 3  # Текущая строка
            if query:
                category = 0
                ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
                ws.page_setup.paperSize = ws.PAPERSIZE_A4
                ws.page_setup.fitToWidth = 1
                ws.column_dimensions['C'].width = 40
                ws.column_dimensions['D'].width = 13
                ws.column_dimensions['E'].width = 13
                ws.column_dimensions['F'].width = 46
                for item in query:
                    row = row + 1
                    if category != item.category_id:
                        category = item.category_id

                        ws.merge_cells('{}{}:{}{}'.format(cols[1], row, cols[8], row))
                        cell = '{}{}'.format(cols[1], row)
                        ws[cell].font = Font(bold=True)
                        ws[cell].fill = PatternFill(fill_type='solid', fgColor='00D0D0D0')
                        ws[cell].value = 'Весовая категория {} {}'.format(
                            MemberCats.objects.get(pk=category).age_category,
                            MemberCats.objects.get(pk=category).weight_category)
                        row = row + 1
                        cell = '{}{}'.format(cols[1], row)
                        ws[cell].value = 'Место'
                        ws[cell].font = Font(bold=True)
                        ws[cell].alignment = Alignment(horizontal='center')
                        cell = '{}{}'.format(cols[2], row)
                        ws[cell].value = 'Фамилия, Имя, Отчество'
                        ws[cell].font = Font(bold=True)
                        ws[cell].alignment = Alignment(horizontal='center')
                        cell = '{}{}'.format(cols[3], row)
                        ws[cell].value = 'Спортивное звание'
                        ws[cell].font = Font(bold=True, size=10)
                        ws[cell].alignment = Alignment(wrap_text=True, horizontal='center')
                        cell = '{}{}'.format(cols[4], row)
                        ws[cell].value = 'Дата рождения'
                        ws[cell].font = Font(bold=True, size=10)
                        ws[cell].alignment = Alignment(wrap_text=True, horizontal='center')
                        cell = '{}{}'.format(cols[5], row)
                        ws[cell].value = 'Субъект Российской Федерации'
                        ws[cell].font = Font(bold=True)
                        ws[cell].alignment = Alignment(horizontal='center')
                        cell = '{}{}'.format(cols[6], row)
                        ws[cell].value = 'Команда'
                        ws[cell].font = Font(bold=True)
                        ws[cell].alignment = Alignment(horizontal='center')
                        cell = '{}{}'.format(cols[7], row)
                        ws[cell].value = 'Очки'
                        ws[cell].font = Font(bold=True)
                        ws[cell].alignment = Alignment(horizontal='center')
                        cell = '{}{}'.format(cols[8], row)
                        ws[cell].value = 'Вес'
                        ws[cell].font = Font(bold=True)
                        ws[cell].alignment = Alignment(horizontal='center')
                        row = row + 1
                    cell = '{}{}'.format(cols[1], row)
                    ws[cell].value = item.place
                    ws[cell].alignment = Alignment(horizontal='center')
                    cell = '{}{}'.format(cols[2], row)
                    member = Members.objects.get(pk=item.member_id)
                    ws[cell].value = '{} {} {}'.format(member.surname, member.name, member.second_name)
                    cell = '{}{}'.format(cols[3], row)
                    ws[cell].value = member.rank
                    ws[cell].alignment = Alignment(horizontal='center')
                    cell = '{}{}'.format(cols[4], row)
                    ws[cell].value = member.birthday
                    ws[cell].alignment = Alignment(horizontal='center')
                    cell = '{}{}'.format(cols[5], row)
                    ws[cell].value = item.fo
                    cell = '{}{}'.format(cols[6], row)
                    ws[cell].value = member.team
                    cell = '{}{}'.format(cols[7], row)
                    ws[cell].value = points[item.place - 1]
                    ws[cell].alignment = Alignment(horizontal='center')
                    cell = '{}{}'.format(cols[8], row)
                    ws[cell].value = member.weight
                    ws[cell].alignment = Alignment(horizontal='center')
                ws.print_area = 'A1:I{}'.format(row)
                set_border(ws, 'B1:I{}'.format(row))
            filename = '{}-{}.xlsx'.format(m1.title, m1.date)
            wb.save(filename)
            try:
                with open(filename, 'rb') as f:
                    file_data = f.read()
                # sending response
                response = HttpResponse(file_data, content_type='application/vnd.ms-excel')
                response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
            except IOError:
                # handle file not exist case here
                response = HttpResponseNotFound('<h1>File not exist</h1>')
            return response



        if action == 'delete_sportsmen':
            if request.user.is_superuser:
                member_id = request.POST.get("member_id", "0")
                if member_id:
                    Members.objects.filter(pk=member_id).delete()
            return redirect(request.path)

        if action == 'delete_member':  # удалить участника из категории
            owner_id = Match.objects.get(pk=match_id).owner_id
            if request.user.id == owner_id:
                member_id = request.POST.get("member_id", "0")
                category_id = request.POST.get("category_id", "0")
                data['match_id'] = match_id
                if member_id:
                    # Members.objects.filter(pk=member_id).delete()
                    StartLists.objects.filter(category_id=category_id, member_id=member_id).delete()
                    count = StartLists.objects.filter(category_id=category_id, step=0).count()
                    MemberCats.objects.filter(pk=category_id).update(member_count=count)
            data['start_list'] = get_list(match_id, request, csrf.get_token(request))
            return redirect('/dashboard/match/?id={}'.format(match_id))

        if action == 'new_category':
            member_id = request.POST.get("member_id", "0")
            category_id = request.POST.get("category_id", "0")
            StartLists.objects.filter(pk=member_id).update(category_id=category_id)
            data['start_list'] = get_list(match_id, request)

        if action == 'to_score':
            return redirect('/dashboard/tables/')

        if action == 'start_match':
            status = Match.objects.get(pk=match_id).status
            if not status:  # Если турнир еще не начался, распределяем категории по столам.
                count = Match.objects.get(pk=match_id).table_count  # Количество столов в турнире
                table = 1
                cats = MemberCats.objects.filter(match_id=match_id).order_by('hand',
                                                                             'weight_category')  # Список категорий
                for cat in cats:  # для каждой категории
                    MemberCats.objects.filter(pk=cat.pk).update(table=table, final=0)  # Определяем стол для категории
                    table = table + 1
                    if table > count:
                        count = 1
                Match.objects.filter(pk=match_id).update(status=1)
            data['start_list'] = get_tables(user_id, match_id, csrf.get_token(request))  # Вывод турнира
            data['match_id'] = match_id

        if action == 'table_list':
            data['start_list'] = get_tables(None, match_id, csrf.get_token(request))  # Вывод турнира
            data['match_id'] = match_id

    if match_id:
        data['match_id'] = match_id
        return render(request, 'pages/match.html', context=data)  #
    else:
        return render(request, 'pages/dashboard.html', context=data)


def one_table(request):
    data = {}
    category_id = request.GET.get("id")
    # data['members'] = get_match_list(match_id=None, category_id=category_id)
    return render(request, 'pages/table.html', context=data)


def select_user(request):
    if request.user.is_authenticated:
        data = {}
        user = request.user
        member_id = request.POST.get("member_id")
        category_id = request.POST.get("category_id")
        match_id = request.POST.get("match_id")
        fo = Members.objects.get(pk=member_id).fo
        order = StartLists.objects.filter(match_id=match_id, category_id=category_id, member_id=member_id, fo=fo)
        if not order:  # если мы не ошиблись и не пытаемся добавить спортсмена еще раз в категорию
            query = StartLists(match_id=match_id, category_id=category_id, member_id=member_id, fo=fo)
            query.save()
            count = StartLists.objects.filter(category_id=category_id, step=0).count()
            MemberCats.objects.filter(pk=category_id).update(member_count=count)
            # Members.objects.filter(pk=member_id).update(weight=0)  # Обнуляем вес участника
        query = Match.objects.get(pk=match_id)
        data['title'] = query.title
        data['date'] = query.date
        data['start_list'] = get_list(match_id, request, csrf.get_token(request))
        data['match_id'] = match_id
        data['header'] = get_header(user)
        data['menu'] = get_menu(request)
        data['menu_cat'] = get_menu_cat(match_id, request)
        return render(request, 'pages/match.html', context=data)
    else:
        return redirect('/')


def show_members(request):
    member_id = None
    if request.method == 'POST':
        member_id = request.POST.get("member_id")
    if request.method == 'GET':
        member_id = request.GET.get("id")
    data = {}
    s = []
    data['menu'] = get_menu(request)
    if not member_id:
        s.append('<table class="tab_group">\n\t<tbody>')
        s.append('<tr class="row_header">')
        s.append('\t<th class="coll_field">Фамилия, Имя, Отчество</th>')
        s.append('\t<th class="coll_field">Дата рождения</th>')
        s.append('\t<th class="coll_field">Спортивное звание</th>')
        s.append('\t<th class="coll_field">Субъект РФ</th>')
        s.append('\t<th class="coll_field"></th>')
        s.append('</tr>')
        items = Members.objects.filter().order_by('surname')
        for member in items:
            s.append('<tr class="row_member">')
            s.append(
                '\t\t\t<td class="coll_name">{} {} {}</td>'.format(member.surname, member.name, member.second_name))
            if member.birthday:
                s.append('\t\t\t<th class="coll_field">{}</th>'.format(member.birthday))
            else:
                s.append('\t\t\t<td class="coll_field"></td>')
            s.append('\t\t\t<th class="coll_field">{}</th>'.format(member.rank))
            s.append('\t\t\t<th class="coll_name">{}</th>'.format(member.fo))

            s.append('\t\t\t<td class="coll_field">'
                     '<form method="post" action="/members/">'
                     '<input type="hidden" name="member_id" value="{}">'.format(member.pk))
            s.append('<input type="hidden" name="csrfmiddlewaretoken" value="{}">'.format(csrf.get_token(request)))
            s.append('<input type="submit" class="btn" id="go_to_match" value="Перейти">')
            s.append('</form>\n\t\t</td>')
            s.append('</tr>')
        s.append('</tbody></table>')
        data['members'] = "\n".join(s)
    else:
        member = Members.objects.get(pk=member_id)
        s.append('<div class=member-item>')
        s.append('<div class="profile"><div class="member_pic">')
        s.append('<img src="/static/media/{}"  class="img-circle member_img">'.format(member.photo))
        s.append('</div></div>')
        s.append(
            '<span class="member-item-name">{} {} {}</span>'.format(member.surname, member.name, member.second_name))
        s.append('<span class="member-item-birthday">дата рождения: {}</span>'.format(member.birthday))
        s.append('<span class="member-item-rank">Спортивное звание: {}</span>'.format(member.rank))
        s.append('<span class="member-item-fo">Субъект РФ: {}</span>'.format(member.fo))
        s.append('<span class="member-item-team">Выступает за команду: {}</span>'.format(member.team))
        s.append('<span class="member-item-tr">Тренируется у: {}</span>'.format(member.trener))
        if request.user.is_superuser:
            s.append(
                '<form method="post" class="member-item-btn" enctype="multipart/form-data" action="/set_user_photo/">'
                '<input type="hidden" name="member_id" value="{}">'.format(member.pk))
            s.append('<input type="hidden" name="csrfmiddlewaretoken" value="{}">'.format(csrf.get_token(request)))
            s.append('<input type="file" name="photo" accept="image/png, image/jpeg" />')
            s.append('<input type="submit" class="btn" id="go_to_match" value="Загрузить фото">')
            s.append('</form>')
            s.append(
                '<form method="post" class="member-item-btn2" enctype="multipart/form-data" action="/dashboard/match/">'
                '<input type="hidden" name="member_id" value="{}">'.format(member.pk))
            s.append('<input type="hidden" name="csrfmiddlewaretoken" value="{}">'.format(csrf.get_token(request)))
            s.append('<input type="hidden" name="selector" value="delete_sportsmen">')
            s.append('<input type="submit" class="btn" id="go_to_match" value="Удалить спортсмена">')
            s.append('</form>')
        s.append('</div>')
        s.append('<table class="tab_group">\n\t<tbody>')
        s.append('<tr class="row_header">')
        s.append('\t<th colspan="3" class="coll_field">Победил спортсменов:</th>')
        s.append('</tr>')
        items = MatchHistory.objects.filter(win_id=member_id)
        for item in items:
            member = Members.objects.get(pk=item.los_id)
            match = Match.objects.get(pk=item.match)
            s.append('<tr class="row_member">')
            s.append(
                '\t\t\t<td class="coll_name">{} {} {}</td>'.format(member.surname, member.name, member.second_name))
            s.append('\t\t\t<th class="coll_min">{}</th>'.format(member.rank))
            s.append('\t\t\t<td class="coll_name">{} {}</td>'.format(match.title, match.date))
            s.append('</tr>')
        s.append('</tbody></table>')
        s.append('<table class="tab_group">\n\t<tbody>')
        s.append('<tr class="row_header">')
        s.append('\t<th colspan="3" class="coll_field">Проиграл спортсменам:</th>')
        s.append('</tr>')
        items = MatchHistory.objects.filter(los_id=member_id)
        for item in items:
            member = Members.objects.get(pk=item.win_id)
            match = Match.objects.get(pk=item.match)
            s.append('<tr class="row_member">')
            s.append(
                '\t\t\t<td class="coll_name">{} {} {}</td>'.format(member.surname, member.name, member.second_name))
            s.append('\t\t\t<th class="coll_min">{}</th>'.format(member.rank))
            s.append('\t\t\t<td class="coll_name">{} {}</td>'.format(match.title, match.date))
            s.append('</tr>')
        s.append('</tbody></table>')
        data['member'] = "\n".join(s)
    return render(request, 'pages/member.html', context=data)


def set_user_photo(request):
    if request.method == 'POST' and request.FILES['photo']:
        file = request.FILES['photo']
        fss = FileSystemStorage()
        new_file = fss.save(file.name, file)
        file_url = fss.url(new_file)
        member_id = request.POST.get('member_id')
        upload_path = settings.MEDIA_ROOT + fss.url(file_url)
        img = Image.open(upload_path)
        width = img.width
        height = img.height
        max_size = max(width, height)
        if max_size > 256:
            output_size = (
                round(width / max_size * 256),  # Сохраняем пропорции
                round(height / max_size * 256))
            img.thumbnail(output_size)
            img.save(upload_path)
        Members.objects.filter(pk=member_id).update(photo=new_file)
    return redirect('/members/?id={}'.format(member_id))


def download_file(request):

    return None